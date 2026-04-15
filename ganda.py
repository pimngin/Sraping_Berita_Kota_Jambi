import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import requests
import cloudscraper
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import quote_plus
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import re
import time

# --- KONFIGURASI & FUNGSI UTILITAS SCRAPER ---
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
}

BATCH_SIZE = 5  # Jumlah halaman yang di-fetch secara paralel

BULAN_ID = {
    "januari": 1,
    "jan": 1,
    "februari": 2,
    "feb": 2,
    "maret": 3,
    "mar": 3,
    "april": 4,
    "apr": 4,
    "mei": 5,
    "juni": 6,
    "jun": 6,
    "juli": 7,
    "jul": 7,
    "agustus": 8,
    "agu": 8,
    "september": 9,
    "sep": 9,
    "oktober": 10,
    "okt": 10,
    "november": 11,
    "nov": 11,
    "desember": 12,
    "des": 12,
}


def clean_text(text):
    if not text:
        return "-"
    return re.sub(r"\s+", " ", text).strip()


def parse_general_date(date_str):
    now = datetime.now()
    date_str = str(date_str).lower().strip().replace("wib", "").replace(",", "")
    if "lalu" in date_str or "baru" in date_str:
        try:
            angka = (
                int(re.search(r"\d+", date_str).group())
                if re.search(r"\d+", date_str)
                else 0
            )
            if "jam" in date_str:
                return now - timedelta(hours=angka)
            elif "menit" in date_str:
                return now - timedelta(minutes=angka)
            elif "hari" in date_str:
                return now - timedelta(days=angka)
            elif "detik" in date_str:
                return now - timedelta(seconds=angka)
        except:
            return now
        return now
    try:
        parts = date_str.split()
        day, month, year = None, None, None
        for i, part in enumerate(parts):
            if part in BULAN_ID:
                month = BULAN_ID[part]
                if i > 0 and parts[i - 1].isdigit():
                    day = int(parts[i - 1])
                if i < len(parts) - 1 and parts[i + 1].isdigit():
                    year = int(parts[i + 1])
                break
        if day and month and year:
            return datetime(year, month, day)
    except:
        pass

    # Format numerik DD-MM-YYYY (contoh: "12-04-2026" dari Jambi Ekspres)
    try:
        match = re.search(r"(\d{1,2})-(\d{1,2})-(\d{4})", date_str)
        if match:
            day, month, year = match.groups()
            return datetime(int(year), int(month), int(day))
    except:
        pass

    return None


def is_in_range(date_obj, start_date, end_date):
    """Mengecek apakah tanggal berita berada di dalam rentang waktu yang diminta."""
    if not date_obj:
        return False
    return start_date <= date_obj <= end_date


def is_older_than_start(date_obj, start_date):
    """Mengecek apakah tanggal berita lebih tua dari batas 'Dari Tanggal' untuk stop scraping."""
    if not date_obj:
        return False
    return date_obj < start_date


def keyword_match(text, keywords):
    """Mengecek apakah teks mengandung SEMUA kata kunci (case-insensitive)."""
    if not keywords:
        return True
    text_lower = text.lower()
    return all(kw.lower() in text_lower for kw in keywords)


def normalize_link(url):
    """Normalisasi URL untuk pengecekan duplikat yang lebih akurat.
    Menghapus trailing slash, query parameter, dan fragment."""
    return url.rstrip("/").split("?")[0].split("#")[0]


def fetch_page(url):
    """Wrapper fetch halaman untuk digunakan bersama ThreadPoolExecutor.
    Mengembalikan (url, response) atau (url, None) jika gagal."""
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        return url, resp
    except:
        return url, None


# --- FUNGSI SCRAPER ---


def scrape_jambikota(start_date, end_date, keywords, status_callback):
    """
    Scrape Pemkot Jambi.
    Situs ini TIDAK punya fitur search server-side, jadi kita scrape semua halaman
    lalu filter berdasarkan keyword secara lokal.
    """
    base_url = "https://jambikota.go.id/informasi/berita"
    page = 1
    scraped_data = []
    status_callback("🔍 Pemkot Jambi: Mencari berita...")
    while True:
        try:
            response = requests.get(
                f"{base_url}?page={page}", headers=HEADERS, timeout=10
            )
            if response.status_code != 200:
                break
            soup = BeautifulSoup(response.text, "html.parser")
            cards = soup.select("#berita-container a.content-card")
            if not cards:
                break
            stop_scraping = False
            for card in cards:
                title = clean_text(
                    card.select_one("h2.card-title").get_text()
                    if card.select_one("h2.card-title")
                    else ""
                )
                desc = clean_text(
                    card.select_one("p").get_text() if card.select_one("p") else ""
                )
                link = card.get("href")
                date_element = card.select_one(".card-actions")
                date_text = clean_text(date_element.get_text() if date_element else "")
                date_obj = parse_general_date(date_text)

                if is_in_range(date_obj, start_date, end_date):
                    # Filter keyword pada judul + deskripsi
                    gabungan = f"{title} {desc}"
                    if keyword_match(gabungan, keywords):
                        scraped_data.append(
                            {
                                "Sumber": "Pemkot Jambi",
                                "Kategori": "Pemerintahan",
                                "Judul": title,
                                "Deskripsi": desc,
                                "Tanggal": date_text,
                                "Link": link,
                            }
                        )
                elif is_older_than_start(date_obj, start_date):
                    stop_scraping = True
                    break
            if stop_scraping:
                break
            page += 1
            status_callback(
                f"🔍 Pemkot Jambi: Halaman {page}... ({len(scraped_data)} ditemukan)"
            )
            time.sleep(1)
        except Exception as e:
            status_callback(f"⚠️ Pemkot Jambi error: {e}")
            break
    status_callback(f"✅ Pemkot Jambi: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def fetch_tribun_article_date(url):
    """
    Fetch halaman detail artikel Tribun untuk mendapatkan tanggal publikasi.
    Selektor sudah diverifikasi dari outerHTML asli:
        div.content-date > time > span
    Mengembalikan (date_text, date_obj) atau ("-", None) jika tidak ditemukan.
    """
    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")

        content_date = soup.find("div", class_="content-date")
        if content_date:
            span = content_date.find("span")
            if span:
                date_text = clean_text(span.get_text())
                date_obj = parse_general_date(date_text)
                if date_obj:
                    return date_text, date_obj
    except:
        pass

    return "-", None


def scrape_tribun_jambi(start_date, end_date, keywords, status_callback):
    """
    Scrape Tribun Jambi - URL: kota-jambi-bahagia/artikel
    Opsi 3: Concurrent fetch (5 sekaligus) + skip fetch jika keyword
    tidak cocok di judul listing (hemat request).
    Format pagination: /?page=2, /?page=3, dst.
    """

    url_base = "https://jambi.tribunnews.com/kota-jambi-bahagia/artikel"
    scraped_data = []
    link_terscrape = set()  # Untuk deteksi duplikat
    page = 1
    status_callback("🔍 Tribun Jambi: Mencari berita...")

    while True:
        try:
            # Format pagination
            if page == 1:
                url = url_base
            else:
                url = f"{url_base}/?page={page}"

            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code != 200:
                break

            soup = BeautifulSoup(response.text, "html.parser")
            articles = soup.find_all("div", class_="artikel-item")
            if not articles:
                break

            # --- Tahap 1: Kumpulkan kandidat artikel dari listing ---
            # Jika ada keyword, cek judul dulu sebelum fetch detail
            # Jika tidak ada keyword, semua artikel masuk sebagai kandidat
            kandidat = []
            for article in articles:
                a_tag = article.find("a")
                if not a_tag:
                    continue

                link = a_tag.get("href", "")
                if not link or link in link_terscrape:
                    continue

                title_el = a_tag.find("div", class_="artikel-title")
                title = clean_text(
                    title_el.find("p").get_text()
                    if title_el and title_el.find("p")
                    else a_tag.get_text()
                )

                kategori_el = a_tag.find("div", class_="artikel-sub")
                kategori = (
                    clean_text(kategori_el.get_text()) if kategori_el else "Berita Kota"
                )

                # Opsi 3: Jika ada keyword, cek judul dulu — jika tidak cocok, skip fetch detail
                if keywords and not keyword_match(title, keywords):
                    continue

                kandidat.append(
                    {
                        "link": link,
                        "title": title,
                        "kategori": kategori,
                    }
                )

            if not kandidat:
                # Tidak ada kandidat di halaman ini sama sekali
                # Tapi belum tentu pagination habis, lanjut ke halaman berikutnya
                page += 1
                time.sleep(1)
                # Batasi maksimal 3 halaman berturut-turut tanpa kandidat
                # untuk mencegah loop tak terbatas saat keyword sangat spesifik
                if page > 50:
                    break
                continue

            # --- Tahap 2: Concurrent fetch detail untuk ambil tanggal ---
            # Fetch 5 artikel sekaligus menggunakan ThreadPoolExecutor
            status_callback(
                f"🔍 Tribun Jambi: Hal. {page} | Fetch tanggal {len(kandidat)} artikel..."
            )

            def fetch_with_meta(item):
                date_text, date_obj = fetch_tribun_article_date(item["link"])
                return {**item, "date_text": date_text, "date_obj": date_obj}

            hasil_fetch = []
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(fetch_with_meta, k): k for k in kandidat}
                for future in as_completed(futures):
                    try:
                        hasil_fetch.append(future.result())
                    except:
                        pass

            # --- Tahap 3: Filter berdasarkan rentang tanggal ---
            stop_scraping = False
            artikel_baru_di_halaman_ini = 0

            # Urutkan hasil fetch berdasarkan urutan asli (as_completed tidak urut)
            hasil_fetch.sort(
                key=lambda x: kandidat.index(
                    next(k for k in kandidat if k["link"] == x["link"])
                )
            )

            for hasil in hasil_fetch:
                link = hasil["link"]
                title = hasil["title"]
                kategori = hasil["kategori"]
                date_text = hasil["date_text"]
                date_obj = hasil["date_obj"]

                if link in link_terscrape:
                    continue

                if is_in_range(date_obj, start_date, end_date):
                    link_terscrape.add(link)
                    artikel_baru_di_halaman_ini += 1
                    scraped_data.append(
                        {
                            "Sumber": "Tribun Jambi",
                            "Kategori": kategori,
                            "Judul": title,
                            "Deskripsi": "-",
                            "Tanggal": date_text,
                            "Link": link,
                        }
                    )
                elif is_older_than_start(date_obj, start_date):
                    stop_scraping = True
                    break

            if stop_scraping:
                break

            if artikel_baru_di_halaman_ini == 0:
                break

            page += 1
            status_callback(
                f"🔍 Tribun Jambi: Halaman {page}... ({len(scraped_data)} ditemukan)"
            )
            time.sleep(1)

        except Exception as e:
            status_callback(f"⚠️ Tribun Jambi error: {e}")
            break

    status_callback(f"✅ Tribun Jambi: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def scrape_jambi_update(start_date, end_date, keywords, status_callback):
    """
    Scrape Jambi Update - URL: https://www.jambiupdate.co/kategori/bisnis
    Pagination: /kategori/bisnis/2, /3, dst.
    Handle 2 struktur HTML yang berbeda:
    - Struktur 1 (halaman awal): div.media → tanggal relatif (span.post-author)
    - Struktur 2 (halaman lanjutan): a.news-list → tanggal absolut (div.author)
    """
    url_base = "https://www.jambiupdate.co/kategori/bisnis"
    scraped_data = []
    link_terscrape = set()  # Untuk deteksi duplikat
    page = 1
    status_callback("🔍 Jambi Update: Mencari berita...")

    while True:
        try:
            # Format pagination
            if page == 1:
                url = url_base
            else:
                url = f"{url_base}/{page}"

            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code != 200:
                break

            soup = BeautifulSoup(response.text, "html.parser")

            stop_scraping = False
            artikel_baru_di_halaman_ini = 0

            # --- Cek Struktur 1: div.media ---
            articles_s1 = soup.find_all("div", class_="media")
            if articles_s1:
                for article in articles_s1:
                    media_body = article.find("div", class_="media-body")
                    if not media_body:
                        continue

                    # Judul & link
                    h3_tag = media_body.find("h3", class_="category-title")
                    if not h3_tag:
                        continue
                    a_tag = h3_tag.find("a")
                    if not a_tag:
                        continue

                    title = clean_text(a_tag.get_text())
                    link = a_tag.get("href", "")
                    if not link or link in link_terscrape:
                        continue

                    # Deskripsi
                    p_tag = media_body.find("p")
                    desc = clean_text(p_tag.get_text()) if p_tag else "-"

                    # Tanggal: span.post-author → "21 Jam yang lalu"
                    date_text = "-"
                    author_info = media_body.find("div", class_="author-info")
                    if author_info:
                        span_date = author_info.find("span", class_="post-author")
                        if span_date:
                            date_text = clean_text(span_date.get_text())

                    date_obj = parse_general_date(date_text)

                    if is_in_range(date_obj, start_date, end_date):
                        link_terscrape.add(link)
                        artikel_baru_di_halaman_ini += 1
                        if keyword_match(f"{title} {desc}", keywords):
                            scraped_data.append(
                                {
                                    "Sumber": "Jambi Update",
                                    "Kategori": "Bisnis",
                                    "Judul": title,
                                    "Deskripsi": desc,
                                    "Tanggal": date_text,
                                    "Link": link,
                                }
                            )
                    elif is_older_than_start(date_obj, start_date):
                        stop_scraping = True
                        break

            # --- Cek Struktur 2: a.news-list ---
            else:
                articles_s2 = soup.find_all("a", class_="news-list")
                if not articles_s2:
                    break

                for article in articles_s2:
                    link = article.get("href", "")
                    if not link or link in link_terscrape:
                        continue

                    # Judul
                    title_el = article.find("div", class_="title")
                    if not title_el:
                        continue
                    title = clean_text(title_el.get_text())

                    # Tanggal: div.author → "Rabu, 04 Februari 2026 - 16:20:18 WIB"
                    date_text = "-"
                    author_el = article.find("div", class_="author")
                    if author_el:
                        date_text = clean_text(author_el.get_text())

                    date_obj = parse_general_date(date_text)

                    # Struktur 2 tidak punya deskripsi
                    desc = "-"

                    if is_in_range(date_obj, start_date, end_date):
                        link_terscrape.add(link)
                        artikel_baru_di_halaman_ini += 1
                        if keyword_match(title, keywords):
                            scraped_data.append(
                                {
                                    "Sumber": "Jambi Update",
                                    "Kategori": "Bisnis",
                                    "Judul": title,
                                    "Deskripsi": desc,
                                    "Tanggal": date_text,
                                    "Link": link,
                                }
                            )
                    elif is_older_than_start(date_obj, start_date):
                        stop_scraping = True
                        break

            if stop_scraping:
                break

            # Stop jika tidak ada artikel baru (pagination habis)
            if artikel_baru_di_halaman_ini == 0:
                break

            page += 1
            status_callback(
                f"🔍 Jambi Update: Halaman {page}... ({len(scraped_data)} ditemukan)"
            )
            time.sleep(1)

        except Exception as e:
            status_callback(f"⚠️ Jambi Update error: {e}")
            break

    status_callback(f"✅ Jambi Update: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def scrape_jambi_one(start_date, end_date, keywords, status_callback):
    """
    Scrape Jambi One - URL: https://www.jambione.com/tag/ekonomi
    Menggunakan cloudscraper untuk bypass Cloudflare.
    Pagination: ?page=2, ?page=3, dst.
    Session cloudscraper di-refresh tiap 10 halaman untuk mencegah token expired.
    """
    url_base = "https://www.jambione.com/tag/ekonomi"
    scraped_data = []
    link_terscrape = set()  # Untuk deteksi duplikat
    page = 1
    status_callback("🔍 Jambi One: Mencari berita...")

    # Inisialisasi cloudscraper
    scraper = cloudscraper.create_scraper(
        browser={"browser": "chrome", "platform": "windows", "mobile": False}
    )

    while True:
        try:
            # Refresh session tiap 10 halaman untuk mencegah token expired
            if page % 10 == 0:
                scraper = cloudscraper.create_scraper(
                    browser={
                        "browser": "chrome",
                        "platform": "windows",
                        "mobile": False,
                    }
                )
                status_callback(f"🔄 Jambi One: Refresh session (halaman {page})...")

            # Format pagination
            if page == 1:
                url = url_base
            else:
                url = f"{url_base}?page={page}"

            response = scraper.get(url, timeout=15)
            if response.status_code != 200:
                break

            soup = BeautifulSoup(response.text, "html.parser")
            articles = soup.find_all("div", class_="latest__item")
            if not articles:
                break

            stop_scraping = False
            artikel_baru_di_halaman_ini = 0

            for article in articles:
                # Judul & link
                judul_el = article.select_one("h2.latest__title a.latest__link")
                if not judul_el:
                    continue

                judul = clean_text(judul_el.get_text())
                link = judul_el.get("href", "")
                if not link or link in link_terscrape:
                    continue

                # Kategori
                kategori_el = article.select_one("h4.latest__subtitle a")
                kategori = (
                    clean_text(kategori_el.get_text()) if kategori_el else "Lainnya"
                )

                # Tanggal: <date class="latest__date">Rabu, 1 April 2026 | 13:12 WIB</date>
                # Strip karakter "|" sebelum di-parse
                date_text = "-"
                date_el = article.select_one("date.latest__date")
                if date_el:
                    date_text = clean_text(date_el.get_text()).replace("|", "").strip()

                date_obj = parse_general_date(date_text)

                if is_in_range(date_obj, start_date, end_date):
                    link_terscrape.add(link)
                    artikel_baru_di_halaman_ini += 1
                    if keyword_match(judul, keywords):
                        scraped_data.append(
                            {
                                "Sumber": "Jambi One",
                                "Kategori": kategori,
                                "Judul": judul,
                                "Deskripsi": "-",
                                "Tanggal": date_text,
                                "Link": link,
                            }
                        )
                elif is_older_than_start(date_obj, start_date):
                    stop_scraping = True
                    break

            if stop_scraping:
                break

            # Stop jika tidak ada artikel baru (pagination habis)
            if artikel_baru_di_halaman_ini == 0:
                break

            page += 1
            status_callback(
                f"🔍 Jambi One: Halaman {page}... ({len(scraped_data)} ditemukan)"
            )
            time.sleep(1)

        except Exception as e:
            status_callback(f"⚠️ Jambi One error: {e}")
            break

    status_callback(f"✅ Jambi One: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def fetch_jambiekspres_date(url):
    """
    Fetch halaman detail artikel Jambi Ekspres untuk mendapatkan tanggal publikasi.
    Selektor sudah diverifikasi dari outerHTML asli:
        div.post-info span.date → "Minggu 12-04-2026,10:07 WIB"
    Mengembalikan (date_text, date_obj) atau ("-", None) jika tidak ditemukan.
    """
    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")

        post_info = soup.find("div", class_="post-info")
        if post_info:
            span_date = post_info.find("span", class_="date")
            if span_date:
                date_text = clean_text(span_date.get_text())
                date_obj = parse_general_date(date_text)
                if date_obj:
                    return date_text, date_obj
    except:
        pass

    return "-", None


def scrape_jambiekspres(start_date, end_date, keywords, status_callback):
    """
    Scrape Jambi Ekspres - URL: https://jambiekspres.disway.id/kategori/ekonomi
    Pagination: /30, /60, /90, dst (kelipatan 30).
    Tanggal diambil dari halaman detail (Opsi 3):
    - Concurrent fetch 5 detail sekaligus
    - Skip fetch detail jika keyword tidak cocok di judul (hemat request)
    """
    url_base = "https://jambiekspres.disway.id/kategori/ekonomi"
    scraped_data = []
    link_terscrape = set()
    page = 1
    status_callback("🔍 Jambi Ekspres: Mencari berita...")

    while True:
        try:
            # Format pagination: kelipatan 30
            if page == 1:
                url = url_base
            else:
                url = f"{url_base}/{(page - 1) * 30}"

            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code != 200:
                break

            soup = BeautifulSoup(response.text, "html.parser")
            articles = soup.find_all("div", class_="media-content")
            if not articles:
                break

            # --- Tahap 1: Kumpulkan kandidat dari listing ---
            kandidat = []
            for article in articles:
                h2_tag = article.find("h2", class_="media-heading")
                if not h2_tag:
                    continue
                a_tag = h2_tag.find("a")
                if not a_tag:
                    continue

                link = a_tag.get("href", "")
                if not link or link in link_terscrape:
                    continue

                title = clean_text(a_tag.get_text())

                kategori_el = article.find("p", class_="text-uppercase")
                kategori = (
                    clean_text(kategori_el.get_text()) if kategori_el else "Ekonomi"
                )

                # Jika ada keyword, cek judul dulu — jika tidak cocok skip fetch detail
                if keywords and not keyword_match(title, keywords):
                    continue

                kandidat.append(
                    {
                        "link": link,
                        "title": title,
                        "kategori": kategori,
                    }
                )

            if not kandidat:
                page += 1
                if page > 50:
                    break
                time.sleep(1)
                continue

            # --- Tahap 2: Concurrent fetch detail untuk ambil tanggal ---
            status_callback(
                f"🔍 Jambi Ekspres: Hal. {page} | Fetch tanggal {len(kandidat)} artikel..."
            )

            def fetch_with_meta(item):
                date_text, date_obj = fetch_jambiekspres_date(item["link"])
                return {**item, "date_text": date_text, "date_obj": date_obj}

            hasil_fetch = []
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(fetch_with_meta, k): k for k in kandidat}
                for future in as_completed(futures):
                    try:
                        hasil_fetch.append(future.result())
                    except:
                        pass

            # --- Tahap 3: Filter berdasarkan rentang tanggal ---
            stop_scraping = False
            artikel_baru_di_halaman_ini = 0

            # Urutkan hasil fetch berdasarkan urutan asli listing
            hasil_fetch.sort(
                key=lambda x: kandidat.index(
                    next(k for k in kandidat if k["link"] == x["link"])
                )
            )

            for hasil in hasil_fetch:
                link = hasil["link"]
                title = hasil["title"]
                kategori = hasil["kategori"]
                date_text = hasil["date_text"]
                date_obj = hasil["date_obj"]

                if link in link_terscrape:
                    continue

                if is_in_range(date_obj, start_date, end_date):
                    link_terscrape.add(link)
                    artikel_baru_di_halaman_ini += 1
                    scraped_data.append(
                        {
                            "Sumber": "Jambi Ekspres",
                            "Kategori": kategori,
                            "Judul": title,
                            "Deskripsi": "-",
                            "Tanggal": date_text,
                            "Link": link,
                        }
                    )
                elif is_older_than_start(date_obj, start_date):
                    stop_scraping = True
                    break

            if stop_scraping:
                break

            if artikel_baru_di_halaman_ini == 0:
                break

            page += 1
            status_callback(
                f"🔍 Jambi Ekspres: Halaman {page}... ({len(scraped_data)} ditemukan)"
            )
            time.sleep(1)

        except Exception as e:
            status_callback(f"⚠️ Jambi Ekspres error: {e}")
            break

    status_callback(f"✅ Jambi Ekspres: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def scrape_antara_jambi(start_date, end_date, keywords, status_callback):
    """
    Scrape Antara News Jambi - versi tim (batch parallel fetch).
    - Jika ada keyword: gunakan fitur search server-side
    - Jika tidak ada keyword: scrape dari halaman /terkini
    Menggunakan ThreadPoolExecutor + BATCH_SIZE untuk fetch paralel.
    """
    scraped_data = []
    seen = set()
    base_domain = "https://jambi.antaranews.com"

    use_search = bool(keywords)
    if use_search:
        query = " ".join(keywords)
        query_url = quote_plus(query)
        status_callback(f"🔍 Antara News: Mencari '{query}'...")
    else:
        status_callback("🔍 Antara News: Mengambil berita terkini...")

    max_pages = 50
    stop_all = False

    def make_url(page):
        if use_search:
            if page == 1:
                return f"{base_domain}/search?q={query_url}"
            return f"{base_domain}/search/{query_url}/{page}"
        else:
            if page == 1:
                return f"{base_domain}/terkini"
            return f"{base_domain}/terkini/{page}"

    def process_page(html):
        results = []
        should_stop = False
        soup = BeautifulSoup(html, "html.parser")

        # Cari semua judul berita di h2/h3 yang ada link ke /berita/
        for h_tag in soup.find_all(["h2", "h3"]):
            a_tag = h_tag.find("a", href=re.compile(r"/berita/\d+/"))
            if not a_tag:
                continue

            judul = clean_text(a_tag.get_text())
            link = a_tag.get("href", "")
            if not link.startswith("http"):
                link = base_domain + link

            nl = normalize_link(link)
            if nl in seen:
                continue

            date_text = "-"
            deskripsi = "-"
            kategori = "Berita"

            # Ambil container utama untuk memudahkan pencarian tanggal & desc
            container = h_tag.find_parent(
                "div",
                class_=re.compile(
                    r"card__post__content|card__post_content|list-article"
                ),
            )
            if not container:
                container = h_tag.find_parent()

            if container:
                # Cari tanggal di span, li, atau time di dekat judul
                for el in container.find_all(["span", "li", "time"]):
                    txt = clean_text(el.get_text())
                    if txt and len(txt) < 80:
                        if (
                            any(b in txt.lower() for b in list(BULAN_ID.keys())[:24])
                            or "lalu" in txt.lower()
                            or "baru saja" in txt.lower()
                        ):
                            date_text = txt
                            break

                # Cari deskripsi dari paragraf yang cukup panjang
                for p_tag in container.find_all("p"):
                    txt = clean_text(p_tag.get_text())
                    if len(txt) > 20 and txt != judul:
                        deskripsi = txt[:200]
                        break

            date_obj = parse_general_date(date_text)

            if is_in_range(date_obj, start_date, end_date):
                gabungan = f"{judul} {deskripsi}"
                if keyword_match(gabungan, keywords):
                    seen.add(nl)
                    results.append(
                        {
                            "Sumber": "Antara News Jambi",
                            "Kategori": kategori,
                            "Judul": judul,
                            "Deskripsi": deskripsi,
                            "Tanggal": date_text,
                            "Link": link,
                        }
                    )
            elif is_older_than_start(date_obj, start_date):
                should_stop = True
                break

        return results, should_stop

    page = 1
    consecutive_empty = 0

    while page <= max_pages and not stop_all:
        # Fetch BATCH_SIZE halaman sekaligus secara paralel
        urls = [
            (make_url(p), p) for p in range(page, min(page + BATCH_SIZE, max_pages + 1))
        ]

        with ThreadPoolExecutor(max_workers=BATCH_SIZE) as executor:
            futures = {}
            for u, p in urls:
                futures[executor.submit(fetch_page, u)] = p
            page_results = {}
            for future in as_completed(futures):
                p = futures[future]
                url_result, resp = future.result()
                page_results[p] = resp

        # Proses berurutan agar logika stop berdasarkan tanggal tetap benar
        for p in range(page, min(page + BATCH_SIZE, max_pages + 1)):
            resp = page_results.get(p)
            if not resp:
                stop_all = True
                break
            items, should_stop = process_page(resp.text)
            scraped_data.extend(items)
            if should_stop:
                stop_all = True
                break
            if not items:
                consecutive_empty += 1
                if consecutive_empty >= 3:
                    stop_all = True
                    break
            else:
                consecutive_empty = 0

        page += BATCH_SIZE
        mode_label = "search" if use_search else "terkini"
        status_callback(
            f"🔍 Antara News: {len(scraped_data)} berita ({mode_label}, hlm {page - 1})"
        )

    status_callback(f"✅ Antara News Jambi: {len(scraped_data)} berita ditemukan.")
    return scraped_data


# --- GUI APPLICATION ---
class AplikasiScraper:
    def __init__(self, root):
        self.root = root
        self.root.title("Scraper Berita Jambi - Pencarian Berdasarkan Topik")
        self.root.geometry("1150x700")
        self.root.configure(padx=10, pady=10)

        self.semua_berita = []
        self.berita_tampil = []

        # --- FRAME ATAS (Judul) ---
        frame_judul = tk.Frame(self.root)
        frame_judul.pack(fill=tk.X, pady=(0, 10))
        tk.Label(
            frame_judul,
            text="🔎 Portal Berita Jambi - Pencarian Topik",
            font=("Arial", 16, "bold"),
        ).pack(side=tk.LEFT)

        # --- FRAME KATA KUNCI ---
        frame_keyword = tk.Frame(self.root)
        frame_keyword.pack(fill=tk.X, pady=(0, 5))

        tk.Label(
            frame_keyword, text="Kata Kunci Berita:", font=("Arial", 10, "bold")
        ).pack(side=tk.LEFT)
        self.entry_keyword = tk.Entry(frame_keyword, width=50, font=("Arial", 11))
        self.entry_keyword.pack(side=tk.LEFT, padx=10)
        self.entry_keyword.insert(0, "")

        tk.Label(
            frame_keyword,
            text="(contoh: banjir, atau: pembangunan jalan)",
            fg="gray",
            font=("Arial", 9),
        ).pack(side=tk.LEFT)

        # --- FRAME FILTER WAKTU & KONTROL ---
        frame_waktu = tk.Frame(self.root)
        frame_waktu.pack(fill=tk.X, pady=(5, 10))

        sekarang = datetime.now()
        bulan_lalu = sekarang.replace(day=1)

        hari_list = [str(i) for i in range(1, 32)]
        bulan_list = [str(i) for i in range(1, 13)]
        tahun_list = [str(i) for i in range(2020, sekarang.year + 2)]

        # --- Bagian "DARI" ---
        tk.Label(frame_waktu, text="Dari:", font=("Arial", 9, "bold")).pack(
            side=tk.LEFT
        )
        self.cb_d_hari = ttk.Combobox(
            frame_waktu, values=hari_list, width=3, state="readonly"
        )
        self.cb_d_hari.set(str(bulan_lalu.day))
        self.cb_d_hari.pack(side=tk.LEFT, padx=2)

        self.cb_d_bulan = ttk.Combobox(
            frame_waktu, values=bulan_list, width=3, state="readonly"
        )
        self.cb_d_bulan.set(str(bulan_lalu.month))
        self.cb_d_bulan.pack(side=tk.LEFT, padx=2)

        self.cb_d_tahun = ttk.Combobox(
            frame_waktu, values=tahun_list, width=5, state="readonly"
        )
        self.cb_d_tahun.set(str(bulan_lalu.year))
        self.cb_d_tahun.pack(side=tk.LEFT, padx=(2, 15))

        # --- Bagian "SAMPAI" ---
        tk.Label(frame_waktu, text="Sampai:", font=("Arial", 9, "bold")).pack(
            side=tk.LEFT
        )
        self.cb_s_hari = ttk.Combobox(
            frame_waktu, values=hari_list, width=3, state="readonly"
        )
        self.cb_s_hari.set(str(sekarang.day))
        self.cb_s_hari.pack(side=tk.LEFT, padx=2)

        self.cb_s_bulan = ttk.Combobox(
            frame_waktu, values=bulan_list, width=3, state="readonly"
        )
        self.cb_s_bulan.set(str(sekarang.month))
        self.cb_s_bulan.pack(side=tk.LEFT, padx=2)

        self.cb_s_tahun = ttk.Combobox(
            frame_waktu, values=tahun_list, width=5, state="readonly"
        )
        self.cb_s_tahun.set(str(sekarang.year))
        self.cb_s_tahun.pack(side=tk.LEFT, padx=(2, 15))

        # Tombol Mulai Scraping
        self.btn_scrape = tk.Button(
            frame_waktu,
            text="🚀 Mulai Cari Berita",
            bg="#4CAF50",
            fg="white",
            font=("Arial", 10, "bold"),
            command=self.mulai_scraping,
        )
        self.btn_scrape.pack(side=tk.LEFT, padx=10)

        # --- FRAME SUMBER (Checkbox) ---
        frame_sumber = tk.Frame(self.root)
        frame_sumber.pack(fill=tk.X, pady=(0, 5))

        tk.Label(frame_sumber, text="Sumber:", font=("Arial", 9, "bold")).pack(
            side=tk.LEFT
        )

        self.var_pemkot = tk.BooleanVar(value=True)
        self.var_tribun = tk.BooleanVar(value=True)
        self.var_jambupdate = tk.BooleanVar(value=True)
        self.var_jambione = tk.BooleanVar(value=True)
        self.var_antara = tk.BooleanVar(value=True)
        self.var_jambiekspres = tk.BooleanVar(value=True)

        tk.Checkbutton(
            frame_sumber, text="Pemkot Jambi", variable=self.var_pemkot
        ).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(
            frame_sumber, text="Tribun Jambi", variable=self.var_tribun
        ).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(
            frame_sumber, text="Jambi Update", variable=self.var_jambupdate
        ).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(frame_sumber, text="Jambi One", variable=self.var_jambione).pack(
            side=tk.LEFT, padx=5
        )
        tk.Checkbutton(
            frame_sumber, text="Antara News Jambi", variable=self.var_antara
        ).pack(side=tk.LEFT, padx=5)
        tk.Checkbutton(
            frame_sumber, text="Jambi Ekspres", variable=self.var_jambiekspres
        ).pack(side=tk.LEFT, padx=5)

        # Status
        self.lbl_status = tk.Label(
            frame_sumber, text="Status: Menunggu instruksi...", fg="blue"
        )
        self.lbl_status.pack(side=tk.RIGHT, padx=5)

        # --- FRAME TENGAH (Filter Tabel & Export) ---
        frame_tengah = tk.Frame(self.root)
        frame_tengah.pack(fill=tk.X, pady=(0, 5))

        tk.Label(frame_tengah, text="Filter Tabel:").pack(side=tk.LEFT)
        self.entry_cari = tk.Entry(frame_tengah, width=40)
        self.entry_cari.pack(side=tk.LEFT, padx=5)
        self.entry_cari.bind("<KeyRelease>", self.cari_data)

        self.lbl_total = tk.Label(
            frame_tengah, text="Total: 0 berita", font=("Arial", 9, "bold"), fg="#333"
        )
        self.lbl_total.pack(side=tk.LEFT, padx=15)

        self.btn_export = tk.Button(
            frame_tengah,
            text="📥 Export Excel",
            bg="#2196F3",
            fg="white",
            font=("Arial", 9, "bold"),
            command=self.export_csv,
        )
        self.btn_export.pack(side=tk.RIGHT)

        # --- FRAME BAWAH (Tabel Data) ---
        frame_bawah = tk.Frame(self.root)
        frame_bawah.pack(fill=tk.BOTH, expand=True)

        kolom = ("Sumber", "Kategori", "Judul", "Deskripsi", "Tanggal", "Link")
        self.tree = ttk.Treeview(frame_bawah, columns=kolom, show="headings")

        for col in kolom:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor=tk.W)
        self.tree.column("Deskripsi", width=250)
        self.tree.column("Judul", width=300)

        scrollbar_y = ttk.Scrollbar(
            frame_bawah, orient=tk.VERTICAL, command=self.tree.yview
        )
        scrollbar_x = ttk.Scrollbar(
            frame_bawah, orient=tk.HORIZONTAL, command=self.tree.xview
        )
        self.tree.configure(
            yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set
        )

        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    def update_status(self, pesan):
        self.root.after(0, lambda: self.lbl_status.config(text=f"Status: {pesan}"))

    def mulai_scraping(self):
        # Ambil kata kunci
        keyword_text = self.entry_keyword.get().strip()
        keywords = keyword_text.split() if keyword_text else []

        if not keywords:
            jawab = messagebox.askyesno(
                "Tanpa Kata Kunci",
                "Anda tidak memasukkan kata kunci pencarian.\n\n"
                "Apakah ingin mengambil SEMUA berita dalam rentang waktu tersebut?\n\n"
                "(Ini mungkin memakan waktu lebih lama)",
            )
            if not jawab:
                return

        # Validasi Tanggal
        try:
            start_date = datetime(
                int(self.cb_d_tahun.get()),
                int(self.cb_d_bulan.get()),
                int(self.cb_d_hari.get()),
            )
            end_date = datetime(
                int(self.cb_s_tahun.get()),
                int(self.cb_s_bulan.get()),
                int(self.cb_s_hari.get()),
                23,
                59,
                59,
            )

            if start_date > end_date:
                messagebox.showerror(
                    "Error Waktu",
                    "'Dari Tanggal' tidak boleh lebih baru dari 'Sampai Tanggal'!",
                )
                return
        except ValueError:
            messagebox.showerror(
                "Error Waktu", "Format tanggal tidak valid! (Contoh salah: 31 Februari)"
            )
            return

        self.btn_scrape.config(state=tk.DISABLED)
        self.semua_berita.clear()
        self.tree.delete(*self.tree.get_children())
        self.lbl_total.config(text="Total: 0 berita")

        thread = threading.Thread(
            target=self.proses_scraping, args=(start_date, end_date, keywords)
        )
        thread.daemon = True
        thread.start()

    def proses_scraping(self, start_date, end_date, keywords):
        kw_text = f"'{' '.join(keywords)}'" if keywords else "semua berita"
        pesan_waktu = (
            f"{start_date.strftime('%d/%m/%Y')} s/d {end_date.strftime('%d/%m/%Y')}"
        )
        self.update_status(f"Mencari {kw_text} | {pesan_waktu}...")

        all_data = []

        threads = []
        results = {}

        def run_scraper(name, func):
            results[name] = func(start_date, end_date, keywords, self.update_status)

        if self.var_pemkot.get():
            t = threading.Thread(target=run_scraper, args=("pemkot", scrape_jambikota))
            threads.append(t)
        if self.var_tribun.get():
            t = threading.Thread(
                target=run_scraper, args=("tribun", scrape_tribun_jambi)
            )
            threads.append(t)
        if self.var_jambupdate.get():
            t = threading.Thread(
                target=run_scraper, args=("jambupdate", scrape_jambi_update)
            )
            threads.append(t)
        if self.var_jambione.get():
            t = threading.Thread(
                target=run_scraper, args=("jambione", scrape_jambi_one)
            )
            threads.append(t)
        if self.var_antara.get():
            t = threading.Thread(
                target=run_scraper, args=("antara", scrape_antara_jambi)
            )
            threads.append(t)
        if self.var_jambiekspres.get():
            t = threading.Thread(
                target=run_scraper, args=("jambiekspres", scrape_jambiekspres)
            )
            threads.append(t)

        for t in threads:
            t.start()

        for t in threads:
            t.join()

        for name, data in results.items():
            all_data.extend(data)

        self.semua_berita = all_data
        self.berita_tampil = self.semua_berita.copy()

        self.root.after(0, self.tampilkan_data)
        self.update_status(
            f"🎉 Selesai! Ditemukan {len(self.semua_berita)} berita untuk {kw_text}."
        )
        self.root.after(0, lambda: self.btn_scrape.config(state=tk.NORMAL))

    def tampilkan_data(self):
        self.tree.delete(*self.tree.get_children())
        for item in self.berita_tampil:
            self.tree.insert(
                "",
                tk.END,
                values=(
                    item["Sumber"],
                    item["Kategori"],
                    item["Judul"],
                    item["Deskripsi"],
                    item["Tanggal"],
                    item["Link"],
                ),
            )
        self.lbl_total.config(text=f"Total: {len(self.berita_tampil)} berita")

    def cari_data(self, event=None):
        query = self.entry_cari.get().lower()
        if not query:
            self.berita_tampil = self.semua_berita.copy()
        else:
            keywords = query.split()
            self.berita_tampil = []
            for baris in self.semua_berita:
                teks_baris = " ".join(str(val) for val in baris.values()).lower()
                if all(kw in teks_baris for kw in keywords):
                    self.berita_tampil.append(baris)

        self.tampilkan_data()

    def export_csv(self):
        if not self.berita_tampil:
            messagebox.showwarning("Perhatian", "Tidak ada data untuk diexport!")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")],
            title="Simpan Data Berita",
            initialfile=f"Data_Berita_{datetime.now().strftime('%d_%m_%Y')}.xlsx",
        )

        if not file_path:
            return

        # Jika user memilih .csv, simpan sebagai CSV biasa
        if file_path.lower().endswith(".csv"):
            try:
                with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                    fieldnames = ["Sumber", "Kategori", "Judul", "Deskripsi", "Tanggal", "Link"]
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(self.berita_tampil)
                messagebox.showinfo("Sukses", f"Data berhasil disimpan di:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Gagal menyimpan file:\n{e}")
            return

        # Export sebagai Excel (.xlsx) dengan lebar kolom otomatis
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Berita"

            # Header
            fieldnames = ["Sumber", "Kategori", "Judul", "Deskripsi", "Tanggal", "Link"]
            header_font = Font(bold=True, size=11)
            for col_idx, header in enumerate(fieldnames, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Data
            for row_idx, item in enumerate(self.berita_tampil, 2):
                for col_idx, key in enumerate(fieldnames, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=item.get(key, "-"))
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Auto-fit lebar kolom berdasarkan isi
            for col_idx, key in enumerate(fieldnames, 1):
                max_length = len(key)  # mulai dari panjang header
                col_letter = get_column_letter(col_idx)
                for row in range(2, ws.max_row + 1):
                    cell_value = str(ws.cell(row=row, column=col_idx).value or "")
                    # Hitung panjang baris terpanjang (untuk wrap_text)
                    max_length = max(max_length, min(len(cell_value), 80))
                # Tambah sedikit padding
                ws.column_dimensions[col_letter].width = max_length + 4

            # Freeze baris header agar tetap terlihat saat scroll
            ws.freeze_panes = "A2"

            wb.save(file_path)
            messagebox.showinfo("Sukses", f"Data berhasil disimpan di:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal menyimpan file:\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = AplikasiScraper(root)
    root.mainloop()
