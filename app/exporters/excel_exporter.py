# app/exporters/excel_exporter.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

FIELDNAMES = ["Sumber", "Kategori", "Judul", "Deskripsi", "Tanggal", "Link"]


def export_excel(data: list, file_path: str):
    """Ekspor list of dict ke file .xlsx dengan formatting otomatis."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Berita"

    # Header
    header_font = Font(bold=True, size=11)
    for col_idx, header in enumerate(FIELDNAMES, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, item in enumerate(data, 2):
        for col_idx, key in enumerate(FIELDNAMES, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=item.get(key, "-"))
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Auto-fit lebar kolom
    for col_idx, key in enumerate(FIELDNAMES, 1):
        max_length = len(key)
        col_letter = get_column_letter(col_idx)
        for row in range(2, ws.max_row + 1):
            cell_value = str(ws.cell(row=row, column=col_idx).value or "")
            max_length = max(max_length, min(len(cell_value), 80))
        ws.column_dimensions[col_letter].width = max_length + 4

    # Freeze baris header
    ws.freeze_panes = "A2"

    wb.save(file_path)
