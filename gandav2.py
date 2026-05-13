import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import requests
import cloudscraper
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import quote_plus, urljoin, urlparse
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import re
import time

# --- KONFIGURASI & FUNGSI UTILITAS SCRAPER ---
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept-Language": "id-ID,id;q=0.9,en-US;q=0.8,en;q=0.7",
}

BATCH_SIZE = 5

BULAN_ID = {
    "januari": 1, "jan": 1, "februari": 2, "feb": 2, "maret": 3, "mar": 3,
    "april": 4, "apr": 4, "mei": 5, "juni": 6, "jun": 6, "juli": 7, "jul": 7,
    "agustus": 8, "agu": 8, "september": 9, "sep": 9, "oktober": 10, "okt": 10,
    "november": 11, "nov": 11, "desember": 12, "des": 12,
    "january": 1, "february": 2, "march": 3, "may": 5, "june": 6, "july": 7,
    "august": 8, "october": 10, "december": 12,
}


def clean_text(text):
    if not text:
        return "-"
    return re.sub(r"\s+", " ", text).strip()


def parse_general_date(date_str):
    now = datetime.now()
    date_str = str(date_str).lower().strip().replace("wib", "").replace(",", "")
    if "lalu" in date_str or "baru" in date_str:
        try:
            angka = int(re.search(r"\d+", date_str).group()) if re.search(r"\d+", date_str) else 0
            if "jam" in date_str: return now - timedelta(hours=angka)
            elif "menit" in date_str: return now - timedelta(minutes=angka)
            elif "hari" in date_str: return now - timedelta(days=angka)
            elif "detik" in date_str: return now - timedelta(seconds=angka)
        except Exception:
            return now
        return now
    try:
        parts = date_str.split()
        day, month, year = None, None, None
        for i, part in enumerate(parts):
            if part in BULAN_ID:
                month = BULAN_ID[part]
                if i > 0 and parts[i - 1].isdigit():
                    day = int(parts[i - 1])
                if not day and i < len(parts) - 1 and parts[i + 1].isdigit() and len(parts[i + 1]) <= 2:
                    day = int(parts[i + 1])
                    if i < len(parts) - 2 and parts[i + 2].isdigit() and len(parts[i + 2]) == 4:
                        year = int(parts[i + 2])
                if not year and i < len(parts) - 1 and parts[i + 1].isdigit() and len(parts[i + 1]) == 4:
                    year = int(parts[i + 1])
                break
        if day and month and year:
            return datetime(year, month, day)
    except Exception:
        pass
    try:
        match = re.search(r"(\d{1,2})-(\d{1,2})-(\d{4})", date_str)
        if match:
            day, month, year = match.groups()
            return datetime(int(year), int(month), int(day))
    except Exception:
        pass
    return None


def is_in_range(date_obj, start_date, end_date):
    if not date_obj: return False
    return start_date <= date_obj <= end_date


def is_older_than_start(date_obj, start_date):
    if not date_obj: return False
    return date_obj < start_date


def keyword_match(text, keywords):
    if not keywords: return True
    text_lower = text.lower()
    return all(kw.lower() in text_lower for kw in keywords)


def normalize_link(url):
    return url.rstrip("/").split("?")[0].split("#")[0]


def fetch_page(url):
    try:
        resp = requests.get(url, headers=HEADERS, timeout=10)
        return url, resp
    except Exception:
        return url, None


# --- FUNGSI SCRAPER ---

def scrape_jambikota(start_date, end_date, keywords, status_callback):
    base_url = "https://jambikota.go.id/informasi/berita"
    page, scraped_data = 1, []
    status_callback("🔍 Pemkot Jambi: Mencari berita...")
    while True:
        try:
            response = requests.get(f"{base_url}?page={page}", headers=HEADERS, timeout=10)
            if response.status_code != 200: break
            soup = BeautifulSoup(response.text, "html.parser")
            cards = soup.select("#berita-container a.content-card")
            if not cards: break
            stop_scraping = False
            for card in cards:
                title = clean_text(card.select_one("h2.card-title").get_text() if card.select_one("h2.card-title") else "")
                desc = clean_text(card.select_one("p").get_text() if card.select_one("p") else "")
                link = card.get("href")
                date_element = card.select_one(".card-actions")
                date_text = clean_text(date_element.get_text() if date_element else "")
                date_obj = parse_general_date(date_text)
                if is_in_range(date_obj, start_date, end_date):
                    if keyword_match(f"{title} {desc}", keywords):
                        scraped_data.append({"Sumber": "Pemkot Jambi", "Kategori": "Pemerintahan", "Judul": title, "Deskripsi": desc, "Tanggal": date_text, "Link": link})
                elif is_older_than_start(date_obj, start_date):
                    stop_scraping = True; break
            if stop_scraping: break
            page += 1
            status_callback(f"🔍 Pemkot Jambi: Halaman {page}... ({len(scraped_data)} ditemukan)")
            time.sleep(1)
        except Exception as e:
            status_callback(f"⚠️ Pemkot Jambi error: {e}"); break
    status_callback(f"✅ Pemkot Jambi: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def fetch_tribun_article_date(url):
    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")
        content_date = soup.find("div", class_="content-date")
        if content_date:
            span = content_date.find("span")
            if span:
                date_text = clean_text(span.get_text())
                date_obj = parse_general_date(date_text)
                if date_obj: return date_text, date_obj
    except Exception:
        pass
    return "-", None


def scrape_tribun_jambi(start_date, end_date, keywords, status_callback):
    url_base = "https://jambi.tribunnews.com/kota-jambi-bahagia/artikel"
    scraped_data, link_terscrape, page = [], set(), 1
    status_callback("🔍 Tribun Jambi: Mencari berita...")
    while True:
        try:
            url = url_base if page == 1 else f"{url_base}/?page={page}"
            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code != 200: break
            soup = BeautifulSoup(response.text, "html.parser")
            articles = soup.find_all("div", class_="artikel-item")
            if not articles: break
            kandidat = []
            for article in articles:
                a_tag = article.find("a")
                if not a_tag: continue
                link = a_tag.get("href", "")
                if not link or link in link_terscrape: continue
                title_el = a_tag.find("div", class_="artikel-title")
                title = clean_text(title_el.find("p").get_text() if title_el and title_el.find("p") else a_tag.get_text())
                kategori_el = a_tag.find("div", class_="artikel-sub")
                kategori = clean_text(kategori_el.get_text()) if kategori_el else "Berita Kota"
                if keywords and not keyword_match(title, keywords): continue
                kandidat.append({"link": link, "title": title, "kategori": kategori})
            if not kandidat:
                page += 1; time.sleep(1)
                if page > 50: break
                continue
            status_callback(f"🔍 Tribun Jambi: Hal. {page} | Fetch tanggal {len(kandidat)} artikel...")
            def fetch_with_meta(item):
                date_text, date_obj = fetch_tribun_article_date(item["link"])
                return {**item, "date_text": date_text, "date_obj": date_obj}
            hasil_fetch = []
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(fetch_with_meta, k): k for k in kandidat}
                for future in as_completed(futures):
                    try: hasil_fetch.append(future.result())
                    except Exception: pass
            stop_scraping, artikel_baru = False, 0
            hasil_fetch.sort(key=lambda x: next((i for i, k in enumerate(kandidat) if k["link"] == x["link"]), 0))
            for hasil in hasil_fetch:
                link, title, kategori = hasil["link"], hasil["title"], hasil["kategori"]
                date_text, date_obj = hasil["date_text"], hasil["date_obj"]
                if link in link_terscrape: continue
                if is_in_range(date_obj, start_date, end_date):
                    link_terscrape.add(link); artikel_baru += 1
                    scraped_data.append({"Sumber": "Tribun Jambi", "Kategori": kategori, "Judul": title, "Deskripsi": "-", "Tanggal": date_text, "Link": link})
                elif is_older_than_start(date_obj, start_date):
                    stop_scraping = True; break
            if stop_scraping or artikel_baru == 0: break
            page += 1
            status_callback(f"🔍 Tribun Jambi: Halaman {page}... ({len(scraped_data)} ditemukan)")
            time.sleep(1)
        except Exception as e:
            status_callback(f"⚠️ Tribun Jambi error: {e}"); break
    status_callback(f"✅ Tribun Jambi: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def scrape_jambi_update(start_date, end_date, keywords, status_callback):
    url_base = "https://www.jambiupdate.co/kategori/bisnis"
    scraped_data, link_terscrape, page = [], set(), 1
    status_callback("🔍 Jambi Update: Mencari berita...")
    while True:
        try:
            url = url_base if page == 1 else f"{url_base}/{page}"
            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code != 200: break
            soup = BeautifulSoup(response.text, "html.parser")
            stop_scraping, artikel_baru = False, 0
            articles_s1 = soup.find_all("div", class_="media")
            if articles_s1:
                for article in articles_s1:
                    media_body = article.find("div", class_="media-body")
                    if not media_body: continue
                    h3_tag = media_body.find("h3", class_="category-title")
                    if not h3_tag: continue
                    a_tag = h3_tag.find("a")
                    if not a_tag: continue
                    title = clean_text(a_tag.get_text())
                    link = a_tag.get("href", "")
                    if not link or link in link_terscrape: continue
                    p_tag = media_body.find("p")
                    desc = clean_text(p_tag.get_text()) if p_tag else "-"
                    date_text = "-"
                    author_info = media_body.find("div", class_="author-info")
                    if author_info:
                        span_date = author_info.find("span", class_="post-author")
                        if span_date: date_text = clean_text(span_date.get_text())
                    date_obj = parse_general_date(date_text)
                    if is_in_range(date_obj, start_date, end_date):
                        link_terscrape.add(link); artikel_baru += 1
                        if keyword_match(f"{title} {desc}", keywords):
                            scraped_data.append({"Sumber": "Jambi Update", "Kategori": "Bisnis", "Judul": title, "Deskripsi": desc, "Tanggal": date_text, "Link": link})
                    elif is_older_than_start(date_obj, start_date):
                        stop_scraping = True; break
            else:
                articles_s2 = soup.find_all("a", class_="news-list")
                if not articles_s2: break
                for article in articles_s2:
                    link = article.get("href", "")
                    if not link or link in link_terscrape: continue
                    title_el = article.find("div", class_="title")
                    if not title_el: continue
                    title = clean_text(title_el.get_text())
                    date_text = "-"
                    author_el = article.find("div", class_="author")
                    if author_el: date_text = clean_text(author_el.get_text())
                    date_obj = parse_general_date(date_text)
                    if is_in_range(date_obj, start_date, end_date):
                        link_terscrape.add(link); artikel_baru += 1
                        if keyword_match(title, keywords):
                            scraped_data.append({"Sumber": "Jambi Update", "Kategori": "Bisnis", "Judul": title, "Deskripsi": "-", "Tanggal": date_text, "Link": link})
                    elif is_older_than_start(date_obj, start_date):
                        stop_scraping = True; break
            if stop_scraping or artikel_baru == 0: break
            page += 1
            status_callback(f"🔍 Jambi Update: Halaman {page}... ({len(scraped_data)} ditemukan)")
            time.sleep(1)
        except Exception as e:
            status_callback(f"⚠️ Jambi Update error: {e}"); break
    status_callback(f"✅ Jambi Update: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def scrape_jambi_one(start_date, end_date, keywords, status_callback):
    url_base = "https://www.jambione.com/tag/ekonomi"
    scraped_data, link_terscrape, page = [], set(), 1
    status_callback("🔍 Jambi One: Mencari berita...")
    scraper = cloudscraper.create_scraper(browser={"browser": "chrome", "platform": "windows", "mobile": False})
    while True:
        try:
            if page % 10 == 0:
                scraper = cloudscraper.create_scraper(browser={"browser": "chrome", "platform": "windows", "mobile": False})
                status_callback(f"🔄 Jambi One: Refresh session (halaman {page})...")
            url = url_base if page == 1 else f"{url_base}?page={page}"
            response = scraper.get(url, timeout=15)
            if response.status_code != 200: break
            soup = BeautifulSoup(response.text, "html.parser")
            articles = soup.find_all("div", class_="latest__item")
            if not articles: break
            stop_scraping, artikel_baru = False, 0
            for article in articles:
                judul_el = article.select_one("h2.latest__title a.latest__link")
                if not judul_el: continue
                judul = clean_text(judul_el.get_text())
                link = judul_el.get("href", "")
                if not link or link in link_terscrape: continue
                kategori_el = article.select_one("h4.latest__subtitle a")
                kategori = clean_text(kategori_el.get_text()) if kategori_el else "Lainnya"
                date_el = article.select_one("date.latest__date")
                date_text = clean_text(date_el.get_text()).replace("|", "").strip() if date_el else "-"
                date_obj = parse_general_date(date_text)
                if is_in_range(date_obj, start_date, end_date):
                    link_terscrape.add(link); artikel_baru += 1
                    if keyword_match(judul, keywords):
                        scraped_data.append({"Sumber": "Jambi One", "Kategori": kategori, "Judul": judul, "Deskripsi": "-", "Tanggal": date_text, "Link": link})
                elif is_older_than_start(date_obj, start_date):
                    stop_scraping = True; break
            if stop_scraping or artikel_baru == 0: break
            page += 1
            status_callback(f"🔍 Jambi One: Halaman {page}... ({len(scraped_data)} ditemukan)")
            time.sleep(1)
        except Exception as e:
            status_callback(f"⚠️ Jambi One error: {e}"); break
    status_callback(f"✅ Jambi One: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def fetch_jambiekspres_date(url):
    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(response.text, "html.parser")
        post_info = soup.find("div", class_="post-info")
        if post_info:
            span_date = post_info.find("span", class_="date")
            if span_date:
                date_text = clean_text(span_date.get_text())
                date_obj = parse_general_date(date_text)
                if date_obj: return date_text, date_obj
    except Exception:
        pass
    return "-", None


def scrape_jambiekspres(start_date, end_date, keywords, status_callback):
    url_base = "https://jambiekspres.disway.id/kategori/ekonomi"
    scraped_data, link_terscrape, page = [], set(), 1
    status_callback("🔍 Jambi Ekspres: Mencari berita...")
    while True:
        try:
            url = url_base if page == 1 else f"{url_base}/{(page - 1) * 30}"
            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code != 200: break
            soup = BeautifulSoup(response.text, "html.parser")
            articles = soup.find_all("div", class_="media-content")
            if not articles: break
            kandidat = []
            for article in articles:
                h2_tag = article.find("h2", class_="media-heading")
                if not h2_tag: continue
                a_tag = h2_tag.find("a")
                if not a_tag: continue
                link = a_tag.get("href", "")
                if not link or link in link_terscrape: continue
                title = clean_text(a_tag.get_text())
                kategori_el = article.find("p", class_="text-uppercase")
                kategori = clean_text(kategori_el.get_text()) if kategori_el else "Ekonomi"
                if keywords and not keyword_match(title, keywords): continue
                kandidat.append({"link": link, "title": title, "kategori": kategori})
            if not kandidat:
                page += 1
                if page > 50: break
                time.sleep(1); continue
            status_callback(f"🔍 Jambi Ekspres: Hal. {page} | Fetch tanggal {len(kandidat)} artikel...")
            def fetch_with_meta(item):
                date_text, date_obj = fetch_jambiekspres_date(item["link"])
                return {**item, "date_text": date_text, "date_obj": date_obj}
            hasil_fetch = []
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(fetch_with_meta, k): k for k in kandidat}
                for future in as_completed(futures):
                    try: hasil_fetch.append(future.result())
                    except Exception: pass
            stop_scraping, artikel_baru = False, 0
            hasil_fetch.sort(key=lambda x: next((i for i, k in enumerate(kandidat) if k["link"] == x["link"]), 0))
            for hasil in hasil_fetch:
                link, title, kategori = hasil["link"], hasil["title"], hasil["kategori"]
                date_text, date_obj = hasil["date_text"], hasil["date_obj"]
                if link in link_terscrape: continue
                if is_in_range(date_obj, start_date, end_date):
                    link_terscrape.add(link); artikel_baru += 1
                    scraped_data.append({"Sumber": "Jambi Ekspres", "Kategori": kategori, "Judul": title, "Deskripsi": "-", "Tanggal": date_text, "Link": link})
                elif is_older_than_start(date_obj, start_date):
                    stop_scraping = True; break
            if stop_scraping or artikel_baru == 0: break
            page += 1
            status_callback(f"🔍 Jambi Ekspres: Halaman {page}... ({len(scraped_data)} ditemukan)")
            time.sleep(1)
        except Exception as e:
            status_callback(f"⚠️ Jambi Ekspres error: {e}"); break
    status_callback(f"✅ Jambi Ekspres: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def scrape_antara_jambi(start_date, end_date, keywords, status_callback):
    scraped_data, seen = [], set()
    base_domain = "https://jambi.antaranews.com"
    use_search = bool(keywords)
    if use_search:
        query = " ".join(keywords)
        query_url = quote_plus(query)
        status_callback(f"🔍 Antara News: Mencari '{query}'...")
    else:
        status_callback("🔍 Antara News: Mengambil berita terkini...")
    max_pages, stop_all = 50, False
    def make_url(page):
        if use_search:
            return f"{base_domain}/search?q={query_url}" if page == 1 else f"{base_domain}/search/{query_url}/{page}"
        else:
            return f"{base_domain}/terkini" if page == 1 else f"{base_domain}/terkini/{page}"
    def process_page(html):
        results, should_stop = [], False
        soup = BeautifulSoup(html, "html.parser")
        for h_tag in soup.find_all(["h2", "h3"]):
            a_tag = h_tag.find("a", href=re.compile(r"/berita/\d+/"))
            if not a_tag: continue
            judul = clean_text(a_tag.get_text())
            link = a_tag.get("href", "")
            if not link.startswith("http"): link = base_domain + link
            nl = normalize_link(link)
            if nl in seen: continue
            date_text, deskripsi, kategori = "-", "-", "Berita"
            container = h_tag.find_parent("div", class_=re.compile(r"card__post__content|card__post_content|list-article"))
            if not container: container = h_tag.find_parent()
            if container:
                for el in container.find_all(["span", "li", "time"]):
                    txt = clean_text(el.get_text())
                    if txt and len(txt) < 80:
                        if any(b in txt.lower() for b in list(BULAN_ID.keys())[:24]) or "lalu" in txt.lower() or "baru saja" in txt.lower():
                            date_text = txt; break
                for p_tag in container.find_all("p"):
                    txt = clean_text(p_tag.get_text())
                    if len(txt) > 20 and txt != judul:
                        deskripsi = txt[:200]; break
            date_obj = parse_general_date(date_text)
            if is_in_range(date_obj, start_date, end_date):
                if keyword_match(f"{judul} {deskripsi}", keywords):
                    seen.add(nl)
                    results.append({"Sumber": "Antara News Jambi", "Kategori": kategori, "Judul": judul, "Deskripsi": deskripsi, "Tanggal": date_text, "Link": link})
            elif is_older_than_start(date_obj, start_date):
                should_stop = True; break
        return results, should_stop
    page, consecutive_empty = 1, 0
    while page <= max_pages and not stop_all:
        urls = [(make_url(p), p) for p in range(page, min(page + BATCH_SIZE, max_pages + 1))]
        with ThreadPoolExecutor(max_workers=BATCH_SIZE) as executor:
            futures = {}
            for u, p in urls: futures[executor.submit(fetch_page, u)] = p
            page_results = {}
            for future in as_completed(futures):
                p = futures[future]
                url_result, resp = future.result()
                page_results[p] = resp
        for p in range(page, min(page + BATCH_SIZE, max_pages + 1)):
            resp = page_results.get(p)
            if not resp: stop_all = True; break
            items, should_stop = process_page(resp.text)
            scraped_data.extend(items)
            if should_stop: stop_all = True; break
            if not items:
                consecutive_empty += 1
                if consecutive_empty >= 3: stop_all = True; break
            else: consecutive_empty = 0
        page += BATCH_SIZE
        status_callback(f"🔍 Antara News: {len(scraped_data)} berita ({'search' if use_search else 'terkini'}, hlm {page - 1})")
    status_callback(f"✅ Antara News Jambi: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def scrape_jambinews(start_date, end_date, keywords, status_callback):
    scraped_data, link_terscrape = [], set()
    url = "https://www.jambinews.id/"
    halaman, max_halaman = 1, 50
    status_callback("🔍 Jambi News: Mencari berita...")
    while url and halaman <= max_halaman:
        try:
            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code != 200: break
            soup = BeautifulSoup(response.text, "html.parser")
            articles = soup.find_all("div", class_="blog-post hentry index-post")
            if not articles: break
            stop_scraping, artikel_baru = False, 0
            for article in articles:
                h2_tag = article.find("h2", class_="post-title")
                if not h2_tag: continue
                a_tag = h2_tag.find("a")
                if not a_tag: continue
                judul = clean_text(a_tag.get_text())
                link = a_tag.get("href", "")
                if not link or link in link_terscrape: continue
                tag_span = article.find("span", class_="post-tag")
                kategori = clean_text(tag_span.get_text()) if tag_span else "Umum"
                date_span = article.find("span", class_="post-date")
                date_text = clean_text(date_span.get_text()) if date_span else "-"
                date_obj = parse_general_date(date_text)
                link_terscrape.add(link); artikel_baru += 1
                if is_older_than_start(date_obj, start_date): stop_scraping = True; break
                if is_in_range(date_obj, start_date, end_date):
                    if keyword_match(judul, keywords):
                        scraped_data.append({"Sumber": "Jambi News", "Kategori": kategori, "Judul": judul, "Deskripsi": "-", "Tanggal": date_text, "Link": link})
            if stop_scraping: break
            if artikel_baru == 0: break
            next_page = soup.find("a", class_="blog-pager-older-link")
            if next_page and next_page.get("href"):
                url = next_page["href"]; halaman += 1
            else: url = None
            status_callback(f"🔍 Jambi News: Halaman {halaman}... ({len(scraped_data)} ditemukan)")
            time.sleep(1)
        except Exception as e:
            status_callback(f"⚠️ Jambi News error: {e}"); break
    status_callback(f"✅ Jambi News: {len(scraped_data)} berita ditemukan.")
    return scraped_data


def fetch_jambilink_date(url):
    try:
        response = requests.get(url, headers=HEADERS, timeout=10)
        if response.status_code != 200: return "-", None
        soup = BeautifulSoup(response.text, "html.parser")
        time_el = soup.find("time", class_="datetime")
        if time_el:
            iso_str = time_el.get("datetime", "")
            if iso_str:
                try:
                    dt = datetime.fromisoformat(iso_str)
                    return dt.strftime("%d/%m/%Y %H:%M") + " WIB", dt.replace(tzinfo=None)
                except ValueError: pass
            text = time_el.get_text().strip()
            if text:
                clean = text.replace("WIB", "").strip()
                try:
                    dt = datetime.strptime(clean, "%d/%m/%Y %H:%M")
                    return text, dt
                except ValueError: return text, None
    except Exception: pass
    return "-", None


def scrape_jambilink(start_date, end_date, keywords, status_callback):
    base_url = "https://www.jambilink.id/indek"
    scraped_data, link_terscrape, halaman, max_halaman = [], set(), 0, 50
    status_callback("🔍 JambiLink: Mencari berita...")
    while halaman < max_halaman:
        try:
            url = f"{base_url}?page={halaman}"
            response = requests.get(url, headers=HEADERS, timeout=10)
            if response.status_code != 200: break
            soup = BeautifulSoup(response.text, "html.parser")
            table = soup.find("table", class_="views-table")
            if not table: break
            tbody = table.find("tbody")
            if not tbody: break
            rows = tbody.find_all("tr")
            if not rows: break
            kandidat = []
            for row in rows:
                td = row.find("td", class_="views-field-title")
                if not td: continue
                first_a = td.find("a")
                if not first_a: continue
                kategori = clean_text(first_a.get_text())
                span_judul = td.find("span", class_="judul")
                if not span_judul: continue
                a_judul = span_judul.find("a")
                if not a_judul: continue
                judul = clean_text(a_judul.get_text())
                link_path = a_judul.get("href", "")
                link = f"https://www.jambilink.id{link_path}" if link_path.startswith("/") else link_path
                if not link or link in link_terscrape: continue
                if keywords and not keyword_match(judul, keywords): continue
                kandidat.append({"link": link, "title": judul, "kategori": kategori})
            if not kandidat:
                halaman += 1
                if halaman > max_halaman: break
                time.sleep(1); continue
            status_callback(f"🔍 JambiLink: Hal. {halaman + 1} | Fetch tanggal {len(kandidat)} artikel...")
            def fetch_with_meta(item):
                date_text, date_obj = fetch_jambilink_date(item["link"])
                return {**item, "date_text": date_text, "date_obj": date_obj}
            hasil_fetch = []
            with ThreadPoolExecutor(max_workers=5) as executor:
                futures = {executor.submit(fetch_with_meta, k): k for k in kandidat}
                for future in as_completed(futures):
                    try: hasil_fetch.append(future.result())
                    except Exception: pass
            stop_scraping, artikel_baru = False, 0
            hasil_fetch.sort(key=lambda x: next((i for i, k in enumerate(kandidat) if k["link"] == x["link"]), 0))
            for hasil in hasil_fetch:
                link, title, kategori = hasil["link"], hasil["title"], hasil["kategori"]
                date_text, date_obj = hasil["date_text"], hasil["date_obj"]
                if link in link_terscrape: continue
                link_terscrape.add(link); artikel_baru += 1
                if is_older_than_start(date_obj, start_date): stop_scraping = True; break
                if is_in_range(date_obj, start_date, end_date):
                    scraped_data.append({"Sumber": "JambiLink", "Kategori": kategori, "Judul": title, "Deskripsi": "-", "Tanggal": date_text, "Link": link})
            if stop_scraping or artikel_baru == 0: break
            next_page = soup.find("li", class_="pager__item--next")
            if next_page and next_page.find("a"): halaman += 1
            else: break
            status_callback(f"🔍 JambiLink: Halaman {halaman + 1}... ({len(scraped_data)} ditemukan)")
            time.sleep(1)
        except Exception as e:
            status_callback(f"⚠️ JambiLink error: {e}"); break
    status_callback(f"✅ JambiLink: {len(scraped_data)} berita ditemukan.")
    return scraped_data


# --- GENERIC / CUSTOM URL SCRAPER ---

def scrape_custom_url(url_target, start_date, end_date, keywords, status_callback, max_pages=20):
    """
    Scraper generik yang bisa mendeteksi artikel dari website apapun.
    Menggunakan heuristik untuk menemukan judul, tanggal, link, dan deskripsi.
    """
    scraped_data, link_terscrape = [], set()
    parsed = urlparse(url_target)
    domain = parsed.netloc or url_target
    base_url = f"{parsed.scheme}://{parsed.netloc}"
    current_url = url_target
    halaman = 1

    status_callback(f"🔍 Custom ({domain}): Menganalisis halaman...")

    while current_url and halaman <= max_pages:
        try:
            try:
                response = requests.get(current_url, headers=HEADERS, timeout=15)
            except Exception:
                scraper_cf = cloudscraper.create_scraper(browser={"browser": "chrome", "platform": "windows", "mobile": False})
                response = scraper_cf.get(current_url, timeout=15)

            if response.status_code != 200:
                status_callback(f"⚠️ Custom ({domain}): HTTP {response.status_code}")
                break

            soup = BeautifulSoup(response.text, "html.parser")
            artikel_baru = 0

            # Strategi: cari semua <a> yang memiliki teks cukup panjang (kemungkinan judul artikel)
            all_links = soup.find_all("a", href=True)
            artikel_links = []

            for a_tag in all_links:
                href = a_tag.get("href", "")
                if not href or href.startswith("#") or href.startswith("javascript"): continue
                # Buat URL absolut
                full_url = urljoin(current_url, href)
                # Filter: hanya link yang masih di domain yang sama
                if urlparse(full_url).netloc != parsed.netloc: continue
                # Cari teks judul
                text = clean_text(a_tag.get_text())
                # Judul artikel biasanya 20-200 karakter
                if len(text) < 20 or len(text) > 300: continue
                # Skip navigasi umum
                skip_words = ["home", "about", "contact", "login", "register", "menu", "search", "privacy", "sitemap"]
                if text.lower().strip() in skip_words: continue
                if full_url in link_terscrape: continue
                artikel_links.append({"url": full_url, "title": text, "element": a_tag})

            if not artikel_links:
                status_callback(f"⚠️ Custom ({domain}): Tidak menemukan artikel di halaman {halaman}")
                break

            status_callback(f"🔍 Custom ({domain}): Hal. {halaman} | Ditemukan {len(artikel_links)} kandidat artikel...")

            for item in artikel_links:
                full_url = item["url"]
                title = item["title"]
                a_tag = item["element"]

                if full_url in link_terscrape: continue

                # Cari tanggal di sekitar link
                date_text = "-"
                date_obj = None
                parent = a_tag.find_parent()
                # Cari di 3 level parent ke atas
                for _ in range(3):
                    if parent is None: break
                    for el in parent.find_all(["time", "span", "div", "p", "small", "date"]):
                        txt = clean_text(el.get_text())
                        if txt and 5 < len(txt) < 80:
                            test_date = parse_general_date(txt)
                            if test_date:
                                date_text = txt
                                date_obj = test_date
                                break
                    if date_obj: break
                    parent = parent.find_parent()

                # Cari deskripsi
                desc = "-"
                parent2 = a_tag.find_parent()
                for _ in range(3):
                    if parent2 is None: break
                    for p_tag in parent2.find_all("p"):
                        txt = clean_text(p_tag.get_text())
                        if len(txt) > 30 and txt != title:
                            desc = txt[:200]
                            break
                    if desc != "-": break
                    parent2 = parent2.find_parent()

                # Cari kategori
                kategori = "Umum"
                parent3 = a_tag.find_parent()
                if parent3:
                    for cat_el in parent3.find_all(["span", "a", "div"], class_=re.compile(r"cat|tag|label|topic|rubrik", re.I)):
                        cat_txt = clean_text(cat_el.get_text())
                        if 2 < len(cat_txt) < 30 and cat_txt != title:
                            kategori = cat_txt
                            break

                # Filter berdasarkan tanggal dan keyword
                in_range = is_in_range(date_obj, start_date, end_date) if date_obj else True
                if in_range:
                    gabungan = f"{title} {desc}"
                    if keyword_match(gabungan, keywords):
                        link_terscrape.add(full_url)
                        artikel_baru += 1
                        scraped_data.append({
                            "Sumber": domain,
                            "Kategori": kategori,
                            "Judul": title,
                            "Deskripsi": desc,
                            "Tanggal": date_text,
                            "Link": full_url,
                        })

            # Cari tombol next page
            next_url = None
            for a_next in soup.find_all("a", href=True):
                txt = clean_text(a_next.get_text()).lower()
                href = a_next.get("href", "")
                if any(nav in txt for nav in ["next", "older", "selanjutnya", "lama", "berikutnya", "»", "›"]):
                    next_url = urljoin(current_url, href)
                    break
                # Cek pattern page=N+1
                if re.search(r"page[=/-]?\d+", href):
                    candidate_url = urljoin(current_url, href)
                    if candidate_url != current_url:
                        next_url = candidate_url

            if artikel_baru == 0 or not next_url:
                break

            current_url = next_url
            halaman += 1
            status_callback(f"🔍 Custom ({domain}): {len(scraped_data)} artikel ditemukan, halaman {halaman}...")
            time.sleep(1.5)

        except Exception as e:
            status_callback(f"⚠️ Custom ({domain}) error: {e}")
            break

    status_callback(f"✅ Custom ({domain}): {len(scraped_data)} artikel ditemukan.")
    return scraped_data


# --- DAFTAR SUMBER TERDAFTAR ---
SUMBER_LIST = [
    ("Pemkot Jambi", scrape_jambikota),
    ("Tribun Jambi", scrape_tribun_jambi),
    ("Jambi Update", scrape_jambi_update),
    ("Jambi One", scrape_jambi_one),
    ("Antara News Jambi", scrape_antara_jambi),
    ("Jambi Ekspres", scrape_jambiekspres),
    ("Jambi News", scrape_jambinews),
    ("JambiLink", scrape_jambilink),
]


# --- GUI APPLICATION ---
class AplikasiScraper:
    def __init__(self, root):
        self.root = root
        self.root.title("Scraper Berita Jambi v2 - Multi Sumber + Custom URL")
        self.root.geometry("1200x750")
        self.root.configure(padx=10, pady=10)
        self.semua_berita = []
        self.berita_tampil = []
        self.sumber_expanded = False  # state buka/tutup panel sumber

        # --- FRAME JUDUL ---
        frame_judul = tk.Frame(self.root)
        frame_judul.pack(fill=tk.X, pady=(0, 10))
        tk.Label(frame_judul, text="🔎 Portal Berita Jambi - Scraper v2",
                 font=("Arial", 16, "bold")).pack(side=tk.LEFT)

        # --- FRAME KEYWORD ---
        frame_keyword = tk.Frame(self.root)
        frame_keyword.pack(fill=tk.X, pady=(0, 5))
        tk.Label(frame_keyword, text="Kata Kunci:",
                 font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        self.entry_keyword = tk.Entry(frame_keyword, width=50, font=("Arial", 11))
        self.entry_keyword.pack(side=tk.LEFT, padx=10)
        tk.Label(frame_keyword, text="(contoh: banjir, atau: pembangunan jalan)",
                 fg="gray", font=("Arial", 9)).pack(side=tk.LEFT)

        # --- FRAME TANGGAL ---
        frame_waktu = tk.Frame(self.root)
        frame_waktu.pack(fill=tk.X, pady=(5, 5))
        sekarang = datetime.now()
        bulan_lalu = sekarang.replace(day=1)
        hari_list = [str(i) for i in range(1, 32)]
        bulan_list = [str(i) for i in range(1, 13)]
        tahun_list = [str(i) for i in range(2020, sekarang.year + 2)]

        tk.Label(frame_waktu, text="Dari:",
                 font=("Arial", 9, "bold")).pack(side=tk.LEFT)
        self.cb_d_hari = ttk.Combobox(frame_waktu, values=hari_list, width=3, state="readonly")
        self.cb_d_hari.set(str(bulan_lalu.day))
        self.cb_d_hari.pack(side=tk.LEFT, padx=2)
        self.cb_d_bulan = ttk.Combobox(frame_waktu, values=bulan_list, width=3, state="readonly")
        self.cb_d_bulan.set(str(bulan_lalu.month))
        self.cb_d_bulan.pack(side=tk.LEFT, padx=2)
        self.cb_d_tahun = ttk.Combobox(frame_waktu, values=tahun_list, width=5, state="readonly")
        self.cb_d_tahun.set(str(bulan_lalu.year))
        self.cb_d_tahun.pack(side=tk.LEFT, padx=(2, 15))

        tk.Label(frame_waktu, text="Sampai:",
                 font=("Arial", 9, "bold")).pack(side=tk.LEFT)
        self.cb_s_hari = ttk.Combobox(frame_waktu, values=hari_list, width=3, state="readonly")
        self.cb_s_hari.set(str(sekarang.day))
        self.cb_s_hari.pack(side=tk.LEFT, padx=2)
        self.cb_s_bulan = ttk.Combobox(frame_waktu, values=bulan_list, width=3, state="readonly")
        self.cb_s_bulan.set(str(sekarang.month))
        self.cb_s_bulan.pack(side=tk.LEFT, padx=2)
        self.cb_s_tahun = ttk.Combobox(frame_waktu, values=tahun_list, width=5, state="readonly")
        self.cb_s_tahun.set(str(sekarang.year))
        self.cb_s_tahun.pack(side=tk.LEFT, padx=(2, 15))

        # --- SUMBER BERITA: Collapsible Checkbox Panel ---
        frame_sumber_header = tk.Frame(self.root)
        frame_sumber_header.pack(fill=tk.X, pady=(5, 0))

        self.btn_toggle = tk.Button(frame_sumber_header, text="▶ Pilih Sumber Berita (klik untuk buka)",
                                    font=("Arial", 10, "bold"), relief=tk.FLAT, cursor="hand2",
                                    command=self.toggle_sumber)
        self.btn_toggle.pack(side=tk.LEFT)

        tk.Button(frame_sumber_header, text="✅ Semua", font=("Arial", 8),
                  command=self.pilih_semua).pack(side=tk.RIGHT, padx=2)
        tk.Button(frame_sumber_header, text="❌ Kosongkan", font=("Arial", 8),
                  command=self.hapus_pilihan).pack(side=tk.RIGHT, padx=2)

        # Panel checkbox (awalnya tersembunyi)
        self.frame_sumber_panel = tk.Frame(self.root, bd=1, relief=tk.GROOVE, padx=10, pady=5)
        # TIDAK di-pack dulu — akan muncul saat toggle

        # Buat checkbutton untuk setiap sumber
        self.sumber_vars = []
        for i, (nama, _) in enumerate(SUMBER_LIST):
            var = tk.BooleanVar(value=True)
            self.sumber_vars.append(var)
            cb = tk.Checkbutton(self.frame_sumber_panel, text=nama, variable=var,
                                font=("Arial", 10), anchor="w")
            # Tata 2 kolom: kolom kiri (i genap), kolom kanan (i ganjil)
            row = i // 2
            col = i % 2
            cb.grid(row=row, column=col, sticky="w", padx=15, pady=2)

        # Tombol mulai scraping sumber terdaftar
        frame_btn_scrape = tk.Frame(self.root)
        frame_btn_scrape.pack(fill=tk.X, pady=(5, 5))

        self.btn_scrape = tk.Button(frame_btn_scrape, text="🚀 Mulai Cari Berita (Sumber Terdaftar)",
                                    bg="#4CAF50", fg="white", font=("Arial", 10, "bold"),
                                    command=self.mulai_scraping)
        self.btn_scrape.pack(side=tk.LEFT)

        # --- CUSTOM URL: Langsung di halaman utama ---
        tk.Frame(self.root, height=1, bg="#ccc").pack(fill=tk.X, pady=5)  # separator

        frame_custom = tk.Frame(self.root)
        frame_custom.pack(fill=tk.X, pady=(0, 5))

        tk.Label(frame_custom, text="🌐 Custom URL:",
                 font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        self.entry_url = tk.Entry(frame_custom, width=55, font=("Arial", 11))
        self.entry_url.pack(side=tk.LEFT, padx=8)
        self.entry_url.insert(0, "https://")

        self.btn_custom = tk.Button(frame_custom, text="🚀 Scrape URL Ini",
                                    bg="#FF9800", fg="white", font=("Arial", 10, "bold"),
                                    command=self.mulai_custom_scraping)
        self.btn_custom.pack(side=tk.LEFT, padx=5)

        tk.Label(frame_custom, text="(scrape web apapun tanpa kode baru)",
                 fg="gray", font=("Arial", 9)).pack(side=tk.LEFT, padx=5)

        # --- STATUS ---
        frame_status = tk.Frame(self.root)
        frame_status.pack(fill=tk.X, pady=(0, 5))
        self.lbl_status = tk.Label(frame_status, text="Status: Menunggu instruksi...",
                                   fg="blue", font=("Arial", 9))
        self.lbl_status.pack(side=tk.LEFT)

        # --- FILTER & EXPORT ---
        frame_tengah = tk.Frame(self.root)
        frame_tengah.pack(fill=tk.X, pady=(0, 5))
        tk.Label(frame_tengah, text="Filter Tabel:").pack(side=tk.LEFT)
        self.entry_cari = tk.Entry(frame_tengah, width=40)
        self.entry_cari.pack(side=tk.LEFT, padx=5)
        self.entry_cari.bind("<KeyRelease>", self.cari_data)
        self.lbl_total = tk.Label(frame_tengah, text="Total: 0 berita",
                                  font=("Arial", 9, "bold"), fg="#333")
        self.lbl_total.pack(side=tk.LEFT, padx=15)
        self.btn_export = tk.Button(frame_tengah, text="📥 Export Excel",
                                    bg="#2196F3", fg="white", font=("Arial", 9, "bold"),
                                    command=self.export_csv)
        self.btn_export.pack(side=tk.RIGHT)

        # --- TABEL DATA ---
        frame_bawah = tk.Frame(self.root)
        frame_bawah.pack(fill=tk.BOTH, expand=True)
        kolom = ("Sumber", "Kategori", "Judul", "Deskripsi", "Tanggal", "Link")
        self.tree = ttk.Treeview(frame_bawah, columns=kolom, show="headings")
        for col in kolom:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=120, anchor=tk.W)
        self.tree.column("Deskripsi", width=250)
        self.tree.column("Judul", width=300)
        sb_y = ttk.Scrollbar(frame_bawah, orient=tk.VERTICAL, command=self.tree.yview)
        sb_x = ttk.Scrollbar(frame_bawah, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        sb_y.pack(side=tk.RIGHT, fill=tk.Y)
        sb_x.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    def toggle_sumber(self):
        """Buka/tutup panel checkbox sumber."""
        if self.sumber_expanded:
            self.frame_sumber_panel.pack_forget()
            self.btn_toggle.config(text="▶ Pilih Sumber Berita (klik untuk buka)")
            self.sumber_expanded = False
        else:
            # Pack setelah header, sebelum tombol scrape
            self.frame_sumber_panel.pack(fill=tk.X, pady=(2, 0),
                                         after=self.btn_toggle.master)
            self.btn_toggle.config(text="▼ Pilih Sumber Berita (klik untuk tutup)")
            self.sumber_expanded = True

    def pilih_semua(self):
        for var in self.sumber_vars:
            var.set(True)

    def hapus_pilihan(self):
        for var in self.sumber_vars:
            var.set(False)

    def update_status(self, pesan):
        self.root.after(0, lambda: self.lbl_status.config(text=f"Status: {pesan}"))

    def get_dates(self):
        try:
            sd = datetime(int(self.cb_d_tahun.get()), int(self.cb_d_bulan.get()),
                          int(self.cb_d_hari.get()))
            ed = datetime(int(self.cb_s_tahun.get()), int(self.cb_s_bulan.get()),
                          int(self.cb_s_hari.get()), 23, 59, 59)
            if sd > ed:
                messagebox.showerror("Error", "'Dari' tidak boleh lebih baru dari 'Sampai'!")
                return None
            return sd, ed
        except ValueError:
            messagebox.showerror("Error", "Format tanggal tidak valid!")
            return None

    def mulai_scraping(self):
        kw_text = self.entry_keyword.get().strip()
        keywords = kw_text.split() if kw_text else []
        if not keywords:
            if not messagebox.askyesno("Tanpa Kata Kunci",
                "Tidak ada kata kunci.\nAmbil SEMUA berita?"): return
        dates = self.get_dates()
        if not dates: return
        start_date, end_date = dates
        # Ambil sumber yang dicentang
        selected = [i for i, var in enumerate(self.sumber_vars) if var.get()]
        if not selected:
            messagebox.showwarning("Perhatian", "Pilih minimal 1 sumber berita!")
            return
        self.btn_scrape.config(state=tk.DISABLED)
        self.btn_custom.config(state=tk.DISABLED)
        self.semua_berita.clear()
        self.tree.delete(*self.tree.get_children())
        self.lbl_total.config(text="Total: 0 berita")
        t = threading.Thread(target=self._run_scraping,
                             args=(start_date, end_date, keywords, selected))
        t.daemon = True
        t.start()

    def _run_scraping(self, start_date, end_date, keywords, sel):
        kw = f"'{' '.join(keywords)}'" if keywords else "semua"
        self.update_status(f"Mencari {kw} | {start_date:%d/%m/%Y} - {end_date:%d/%m/%Y}...")
        threads, results = [], {}
        def run(name, func):
            results[name] = func(start_date, end_date, keywords, self.update_status)
        for idx in sel:
            nama, func = SUMBER_LIST[idx]
            threads.append(threading.Thread(target=run, args=(nama, func)))
        for t in threads: t.start()
        for t in threads: t.join()
        all_data = []
        for data in results.values(): all_data.extend(data)
        self.semua_berita = all_data
        self.berita_tampil = all_data.copy()
        self.root.after(0, self.tampilkan_data)
        self.update_status(f"🎉 Selesai! {len(all_data)} berita ditemukan.")
        self.root.after(0, lambda: self.btn_scrape.config(state=tk.NORMAL))
        self.root.after(0, lambda: self.btn_custom.config(state=tk.NORMAL))

    def mulai_custom_scraping(self):
        url = self.entry_url.get().strip()
        if not url or url == "https://":
            messagebox.showwarning("Perhatian", "Masukkan URL target!")
            return
        if not url.startswith("http"): url = "https://" + url
        kw_text = self.entry_keyword.get().strip()
        keywords = kw_text.split() if kw_text else []
        dates = self.get_dates()
        if not dates: return
        start_date, end_date = dates
        self.btn_scrape.config(state=tk.DISABLED)
        self.btn_custom.config(state=tk.DISABLED)
        self.semua_berita.clear()
        self.tree.delete(*self.tree.get_children())
        self.lbl_total.config(text="Total: 0 berita")
        t = threading.Thread(target=self._run_custom,
                             args=(url, start_date, end_date, keywords))
        t.daemon = True
        t.start()

    def _run_custom(self, url, start_date, end_date, keywords):
        self.update_status(f"Scraping: {url}...")
        data = scrape_custom_url(url, start_date, end_date, keywords, self.update_status)
        self.semua_berita = data
        self.berita_tampil = data.copy()
        self.root.after(0, self.tampilkan_data)
        self.update_status(f"🎉 Selesai! {len(data)} artikel dari {url}")
        self.root.after(0, lambda: self.btn_scrape.config(state=tk.NORMAL))
        self.root.after(0, lambda: self.btn_custom.config(state=tk.NORMAL))

    def tampilkan_data(self):
        self.tree.delete(*self.tree.get_children())
        for item in self.berita_tampil:
            self.tree.insert("", tk.END, values=(
                item["Sumber"], item["Kategori"], item["Judul"],
                item["Deskripsi"], item["Tanggal"], item["Link"]))
        self.lbl_total.config(text=f"Total: {len(self.berita_tampil)} berita")

    def cari_data(self, event=None):
        q = self.entry_cari.get().lower()
        if not q:
            self.berita_tampil = self.semua_berita.copy()
        else:
            kws = q.split()
            self.berita_tampil = [
                b for b in self.semua_berita
                if all(k in " ".join(str(v) for v in b.values()).lower() for k in kws)]
        self.tampilkan_data()

    def export_csv(self):
        if not self.berita_tampil:
            messagebox.showwarning("Perhatian", "Tidak ada data untuk diexport!")
            return
        fp = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")],
            title="Simpan Data Berita",
            initialfile=f"Data_Berita_{datetime.now().strftime('%d_%m_%Y')}.xlsx")
        if not fp: return

        fnames = ["Sumber", "Kategori", "Judul", "Deskripsi", "Tanggal", "Link"]
        if fp.lower().endswith(".csv"):
            try:
                with open(fp, "w", newline="", encoding="utf-8-sig") as f:
                    w = csv.DictWriter(f, fieldnames=fnames)
                    w.writeheader()
                    w.writerows(self.berita_tampil)
                messagebox.showinfo("Sukses", f"Tersimpan: {fp}")
            except Exception as e:
                messagebox.showerror("Error", f"Gagal: {e}")
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data Berita"
            hf = Font(bold=True, size=11)
            for ci, h in enumerate(fnames, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = hf
                c.alignment = Alignment(horizontal="center")
            for ri, item in enumerate(self.berita_tampil, 2):
                for ci, key in enumerate(fnames, 1):
                    c = ws.cell(row=ri, column=ci, value=item.get(key, "-"))
                    c.alignment = Alignment(vertical="top", wrap_text=True)
            for ci, key in enumerate(fnames, 1):
                ml = len(key)
                cl = get_column_letter(ci)
                for r in range(2, ws.max_row + 1):
                    cv = str(ws.cell(row=r, column=ci).value or "")
                    ml = max(ml, min(len(cv), 80))
                ws.column_dimensions[cl].width = ml + 4
            ws.freeze_panes = "A2"
            wb.save(fp)
            messagebox.showinfo("Sukses", f"Tersimpan: {fp}")
        except Exception as e:
            messagebox.showerror("Error", f"Gagal: {e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = AplikasiScraper(root)
    root.mainloop()

